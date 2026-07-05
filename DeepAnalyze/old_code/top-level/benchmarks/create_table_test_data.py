#!/usr/bin/env python3
"""
Table Test Data Generator + KB Creation + Upload
=================================================
Generates diverse test table files, creates a knowledge base,
and uploads all files for end-to-end table processing testing.

Usage:
    python3 benchmarks/create_table_test_data.py              # Generate files only
    python3 benchmarks/create_table_test_data.py --upload      # Generate + create KB + upload
    python3 benchmarks/create_table_test_data.py --upload-only # Skip generation, only upload
"""

import argparse
import csv
import io
import json
import os
import shutil
import sys
import time
import urllib.request
import urllib.error
from datetime import datetime, timedelta
from pathlib import Path
from random import Random

try:
    import pandas as pd
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    print("ERROR: pandas and openpyxl are required: pip install pandas openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BASE_URL = os.environ.get("DEEPANALYZE_URL", "http://localhost:21000")
OUTPUT_DIR = Path(__file__).parent / "test-data" / "tables"
EXISTING_EXCEL = Path("/mnt/d/testdata/execl/athlete_events.xlsx")

rand = Random(42)  # Fixed seed for reproducibility


# ============================================================================
# PART 1: Test Data Generation
# ============================================================================

def generate_employee_basic() -> Path:
    """50 rows × 8 columns: basic CSV with mixed Chinese/English columns and data types."""
    path = OUTPUT_DIR / "employee_basic.csv"
    departments = ["技术部", "市场部", "财务部", "人事部", "运营部", "产品部"]
    positions = ["工程师", "经理", "分析师", "设计师", "总监", "专员", "主管"]
    cities = ["北京", "上海", "深圳", "杭州", "广州", "成都"]

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["员工ID", "name", "部门", "age", "salary", "入职日期", "city", "performance_score"])
        for i in range(1, 51):
            writer.writerow([
                f"EMP{i:04d}",
                f"Employee_{i}",
                rand.choice(departments),
                rand.randint(22, 55),
                round(rand.uniform(8000, 50000), 2),
                (datetime(2020, 1, 1) + timedelta(days=rand.randint(0, 1825))).strftime("%Y-%m-%d"),
                rand.choice(cities),
                round(rand.uniform(60, 100), 1),
            ])
    return path


def generate_sales_monthly() -> Path:
    """120 rows × 6 columns: dates + amounts + categories, testing number precision."""
    path = OUTPUT_DIR / "sales_monthly.csv"
    categories = ["电子产品", "服装", "食品", "家居", "图书", "运动"]
    regions = ["华东", "华南", "华北", "西南", "东北"]

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["日期", "category", "region", "revenue", "quantity", "unit_price"])
        base_date = datetime(2024, 1, 1)
        for i in range(120):
            date = base_date + timedelta(days=i * 3)
            revenue = round(rand.uniform(1000.50, 99999.99), 2)
            qty = rand.randint(1, 500)
            writer.writerow([
                date.strftime("%Y-%m-%d"),
                rand.choice(categories),
                rand.choice(regions),
                revenue,
                qty,
                round(revenue / qty, 2),
            ])
    return path


def generate_products() -> Path:
    """30 rows × 5 columns: CSV with commas and quotes (robustness test)."""
    path = OUTPUT_DIR / "products.csv"
    # Some product names contain commas and quotes
    product_names = [
        'T-Shirt, "Premium Quality"',
        'Laptop, 15" Screen',
        'Coffee, "Arabica" Blend',
        'Book: "The Art of Programming"',
        'Mouse, Wireless, "Ergonomic"',
    ]

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["product_id", "product_name", "category", "price", "stock"])
        for i in range(1, 31):
            name = rand.choice(product_names) if i % 3 == 0 else f"Product Item {i}"
            writer.writerow([
                f"SKU{i:05d}",
                name,
                rand.choice(["Electronics", "Clothing", "Food", "Books", "Home"]),
                round(rand.uniform(10, 500), 2),
                rand.randint(0, 1000),
            ])
    return path


def generate_inventory_xlsx() -> Path:
    """200 rows × 7 columns: basic single-sheet Excel."""
    path = OUTPUT_DIR / "inventory.xlsx"
    categories = ["电脑", "手机", "平板", "耳机", "键盘", "鼠标", "显示器", "打印机"]
    statuses = ["在库", "已售", "预定", "退货"]

    wb = Workbook()
    ws = wb.active
    ws.title = "库存明细"
    ws.append(["商品编号", "商品名称", "category", "数量", "单价", "总价值", "状态"])
    for i in range(1, 201):
        qty = rand.randint(0, 500)
        price = round(rand.uniform(50, 15000), 2)
        ws.append([
            f"INV{i:06d}",
            f"商品_{i}",
            rand.choice(categories),
            qty,
            price,
            round(qty * price, 2),
            rand.choice(statuses),
        ])
    wb.save(path)
    return path


def generate_financial_report_xlsx() -> Path:
    """Multi-sheet XLSX: 资产负债表 / 利润表 / 现金流量表."""
    path = OUTPUT_DIR / "financial_report.xlsx"
    wb = Workbook()

    # Sheet 1: 资产负债表 (Balance Sheet) - 60 rows
    ws1 = wb.active
    ws1.title = "资产负债表"
    ws1.append(["科目", "期初余额", "期末余额", "变动金额", "变动比例"])
    items_bs = ["货币资金", "应收账款", "存货", "固定资产", "无形资产", "短期借款",
                "应付账款", "长期借款", "实收资本", "资本公积", "盈余公积", "未分配利润"]
    for _ in range(60):
        opening = round(rand.uniform(100000, 5000000), 2)
        closing = round(opening * rand.uniform(0.8, 1.2), 2)
        ws1.append([
            rand.choice(items_bs),
            opening,
            closing,
            round(closing - opening, 2),
            round((closing - opening) / opening * 100, 2),
        ])

    # Sheet 2: 利润表 (Income Statement) - 80 rows
    ws2 = wb.create_sheet("利润表")
    ws2.append(["月份", "营业收入", "营业成本", "毛利润", "营业费用", "净利润"])
    for month in range(1, 81):
        revenue = round(rand.uniform(500000, 2000000), 2)
        cost = round(revenue * rand.uniform(0.4, 0.7), 2)
        gross = round(revenue - cost, 2)
        expense = round(revenue * rand.uniform(0.1, 0.3), 2)
        net = round(gross - expense, 2)
        ws2.append([f"2024-{((month - 1) % 12) + 1:02d}", revenue, cost, gross, expense, net])

    # Sheet 3: 现金流量表 (Cash Flow) - 50 rows
    ws3 = wb.create_sheet("现金流量表")
    ws3.append(["项目", "本期金额", "上期金额", "同比增减"])
    items_cf = ["经营活动现金流入", "经营活动现金流出", "投资活动现金流入",
                "投资活动现金流出", "筹资活动现金流入", "筹资活动现金流出"]
    for _ in range(50):
        current = round(rand.uniform(100000, 3000000), 2)
        previous = round(current * rand.uniform(0.7, 1.3), 2)
        ws3.append([rand.choice(items_cf), current, previous, round(current - previous, 2)])

    wb.save(path)
    return path


def generate_department_hierarchy_xlsx() -> Path:
    """Multi-sheet XLSX with cross-sheet foreign keys: 员工表/部门表/项目表."""
    path = OUTPUT_DIR / "department_hierarchy.xlsx"
    wb = Workbook()

    # Sheet 1: 部门表 (Departments)
    ws_dept = wb.active
    ws_dept.title = "部门表"
    ws_dept.append(["dept_id", "dept_name", "manager", "budget", "location"])
    departments = [
        ("D001", "技术研发部", "张伟", 5000000, "A栋3楼"),
        ("D002", "市场营销部", "李娜", 3000000, "B栋5楼"),
        ("D003", "财务部", "王强", 2000000, "A栋2楼"),
        ("D004", "人力资源部", "赵敏", 1500000, "A栋1楼"),
        ("D005", "产品设计部", "陈浩", 2500000, "C栋4楼"),
        ("D006", "运维部", "刘洋", 1800000, "A栋负1楼"),
    ]
    for dept in departments:
        ws_dept.append(list(dept))

    # Sheet 2: 员工表 (Employees) - references dept_id
    ws_emp = wb.create_sheet("员工表")
    ws_emp.append(["employee_id", "name", "dept_id", "position", "salary", "hire_date"])
    positions_by_dept = {
        "D001": ["高级工程师", "架构师", "前端工程师", "后端工程师", "测试工程师"],
        "D002": ["市场专员", "品牌经理", "渠道经理", "活动策划"],
        "D003": ["会计", "出纳", "审计", "财务分析师"],
        "D004": ["招聘专员", "培训主管", "薪酬专员", "HRBP"],
        "D005": ["产品经理", "UI设计师", "UX研究员", "交互设计师"],
        "D006": ["运维工程师", "DBA", "网络工程师", "安全工程师"],
    }
    for i in range(1, 51):
        dept_id = rand.choice(["D001", "D002", "D003", "D004", "D005", "D006"])
        ws_emp.append([
            f"E{i:04d}",
            f"员工_{i}",
            dept_id,
            rand.choice(positions_by_dept[dept_id]),
            round(rand.uniform(8000, 45000), 2),
            (datetime(2019, 1, 1) + timedelta(days=rand.randint(0, 2000))).strftime("%Y-%m-%d"),
        ])

    # Sheet 3: 项目表 (Projects) - references dept_id
    ws_proj = wb.create_sheet("项目表")
    ws_proj.append(["project_id", "project_name", "dept_id", "start_date", "end_date", "status", "budget"])
    statuses = ["进行中", "已完成", "已暂停", "规划中"]
    for i in range(1, 21):
        dept_id = rand.choice(["D001", "D002", "D005", "D006"])
        start = datetime(2023, 1, 1) + timedelta(days=rand.randint(0, 500))
        end = start + timedelta(days=rand.randint(30, 365))
        ws_proj.append([
            f"P{i:03d}",
            f"项目_{chr(64 + i) if i <= 26 else 'Z'}{i}",
            dept_id,
            start.strftime("%Y-%m-%d"),
            end.strftime("%Y-%m-%d"),
            rand.choice(statuses),
            round(rand.uniform(100000, 2000000), 2),
        ])

    wb.save(path)
    return path


def generate_sensor_data() -> Path:
    """10,000 rows × 5 columns: large CSV for performance testing."""
    path = OUTPUT_DIR / "sensor_data.csv"
    sensor_ids = ["SENSOR_001", "SENSOR_002", "SENSOR_003", "SENSOR_004", "SENSOR_005"]

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["timestamp", "sensor_id", "temperature", "humidity", "pressure"])
        base_time = datetime(2024, 1, 1, 0, 0, 0)
        for i in range(10000):
            ts = base_time + timedelta(minutes=i * 5)
            writer.writerow([
                ts.strftime("%Y-%m-%d %H:%M:%S"),
                rand.choice(sensor_ids),
                round(rand.uniform(15.0, 45.0), 2),
                round(rand.uniform(20.0, 95.0), 2),
                round(rand.uniform(990.0, 1030.0), 2),
            ])
    return path


def generate_empty_table() -> Path:
    """Empty CSV with only headers."""
    path = OUTPUT_DIR / "empty_table.csv"
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "name", "value"])
    return path


def generate_single_row() -> Path:
    """CSV with exactly 1 data row."""
    path = OUTPUT_DIR / "single_row.csv"
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "name", "status", "amount"])
        writer.writerow(["1", "唯一记录", "active", "999.99"])
    return path


def generate_wide_table() -> Path:
    """20 rows × 30 columns: wide table."""
    path = OUTPUT_DIR / "wide_table.csv"
    headers = [f"col_{i:02d}" for i in range(1, 31)]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row_idx in range(1, 21):
            row = [f"R{row_idx}_C{i}" for i in range(1, 31)]
            # Make some columns numeric
            for j in [3, 5, 8, 12, 15, 20, 25, 28]:
                row[j - 1] = str(round(rand.uniform(0, 100), 2))
            writer.writerow(row)
    return path


def generate_tab_separated() -> Path:
    """TSV format (tab-delimited .csv) - 50 rows × 6 columns."""
    path = OUTPUT_DIR / "tab_separated.csv"
    with open(path, "w", newline="", encoding="utf-8") as f:
        f.write("日期\t城市\t温度\t湿度\t天气\t风速\n")
        cities = ["北京", "上海", "广州", "深圳", "成都", "杭州"]
        weathers = ["晴", "多云", "阴", "小雨", "大雨", "雷阵雨"]
        for i in range(50):
            date = datetime(2024, 1, 1) + timedelta(days=i)
            f.write(f"{date.strftime('%Y-%m-%d')}\t{rand.choice(cities)}\t"
                    f"{round(rand.uniform(-5, 38), 1)}\t{round(rand.uniform(30, 95), 1)}\t"
                    f"{rand.choice(weathers)}\t{round(rand.uniform(0, 30), 1)}\n")
    return path


def generate_unicode_data() -> Path:
    """30 rows × 4 columns: multi-language content (Chinese, English, Japanese, Korean, Arabic)."""
    path = OUTPUT_DIR / "unicode_data.csv"
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "chinese", "english", "multilingual"])
        multilingual_samples = [
            "こんにちは世界",  # Japanese
            "안녕하세요 세계",  # Korean
            "مرحبا بالعالم",   # Arabic
            "Привет мир",      # Russian
            "Bonjour le monde", # French
            "Hallo Welt",      # German
            "¡Hola Mundo!",    # Spanish
            "Olá Mundo",       # Portuguese
            "Ciao Mondo",      # Italian
        ]
        for i in range(1, 31):
            writer.writerow([
                i,
                f"中文内容_{i}号：包含各种标点符号，如逗号、句号、感叹号！",
                f"English content #{i}: includes punctuation, numbers ({rand.randint(100, 999)}), and symbols!",
                rand.choice(multilingual_samples),
            ])
    return path


def generate_crossref_orders_xlsx() -> Path:
    """500 rows × 8 columns: orders table referencing employee IDs from department_hierarchy.xlsx."""
    path = OUTPUT_DIR / "crossref_orders.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "订单表"

    ws.append(["order_id", "employee_id", "customer_name", "product", "quantity", "unit_price", "total", "order_date"])
    products = ["笔记本电脑", "智能手机", "平板电脑", "无线耳机", "机械键盘", "显示器", "打印机", "路由器"]

    for i in range(1, 501):
        employee_id = f"E{rand.randint(1, 50):04d}"  # References department_hierarchy.xlsx employee IDs
        qty = rand.randint(1, 20)
        price = round(rand.uniform(100, 15000), 2)
        ws.append([
            f"ORD{i:06d}",
            employee_id,
            f"客户_{rand.randint(1, 200)}",
            rand.choice(products),
            qty,
            price,
            round(qty * price, 2),
            (datetime(2024, 1, 1) + timedelta(days=rand.randint(0, 365))).strftime("%Y-%m-%d"),
        ])

    wb.save(path)
    return path


# ============================================================================
# PART 2: KB Creation + Upload
# ============================================================================

def api_request(method: str, path: str, data: dict | None = None, timeout: int = 30) -> dict:
    """Make an API request and return parsed JSON response."""
    url = f"{BASE_URL}{path}"
    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, method=method)
    req.add_header("Content-Type", "application/json")
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        error_body = e.read().decode() if e.fp else ""
        return {"error": f"HTTP {e.code}: {error_body[:500]}"}
    except Exception as e:
        return {"error": str(e)}


def upload_file(kb_id: str, file_path: Path, folder: str = "") -> dict:
    """Upload a single file to a KB using multipart form data."""
    url = f"{BASE_URL}/api/knowledge/kbs/{kb_id}/upload"
    filename = file_path.name

    # Build multipart form data
    boundary = "----FormBoundary7MA4YWxkTrZu0gW"
    with open(file_path, "rb") as f:
        file_content = f.read()

    parts = []
    # folder field
    parts.append(f"--{boundary}\r\n".encode())
    parts.append(f'Content-Disposition: form-data; name="folder"\r\n\r\n'.encode())
    parts.append(f"{folder}\r\n".encode())
    # file field
    parts.append(f"--{boundary}\r\n".encode())
    parts.append(
        f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'.encode()
    )
    parts.append(b"Content-Type: application/octet-stream\r\n\r\n")
    parts.append(file_content)
    parts.append(f"\r\n--{boundary}--\r\n".encode())

    body = b"".join(parts)
    req = urllib.request.Request(url, data=body, method="POST")
    req.add_header("Content-Type", f"multipart/form-data; boundary={boundary}")

    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        error_body = e.read().decode() if e.fp else ""
        return {"error": f"HTTP {e.code}: {error_body[:500]}"}
    except Exception as e:
        return {"error": str(e)}


def wait_for_documents_ready(kb_id: str, expected_count: int, timeout: int = 600) -> list[dict]:
    """Poll until all documents in KB reach status=ready or error."""
    start_time = time.time()
    print(f"  Waiting for {expected_count} documents to be processed...")

    while time.time() - start_time < timeout:
        result = api_request("GET", f"/api/knowledge/kbs/{kb_id}/documents")
        if "error" in result:
            print(f"  Poll error: {result['error']}")
            time.sleep(5)
            continue

        docs = result if isinstance(result, list) else result.get("documents", result.get("data", []))
        if not isinstance(docs, list):
            docs = []

        ready_count = sum(1 for d in docs if d.get("status") == "ready")
        error_count = sum(1 for d in docs if d.get("status") == "error")
        total = len(docs)

        if total >= expected_count:
            if ready_count + error_count >= expected_count:
                print(f"  All documents processed: {ready_count} ready, {error_count} errors")
                return docs

        elapsed = time.time() - start_time
        print(f"  [{elapsed:.0f}s] {ready_count}/{expected_count} ready, {error_count} errors")
        time.sleep(5)

    print(f"  TIMEOUT after {timeout}s")
    return []


def create_kb_and_upload(table_files: list[Path], copy_existing_excel: bool = True) -> dict:
    """Create a knowledge base and upload all table test files."""
    # 1. Create KB
    print("\n[1/3] Creating knowledge base...")
    kb_result = api_request("POST", "/api/knowledge/kbs", {
        "name": "表格处理测试库",
        "description": "NativeTableProcessor 全链路测试 - 包含多种格式的表格文件",
        "visibility": "private",
    })

    if "error" in kb_result:
        print(f"ERROR: Failed to create KB: {kb_result['error']}")
        return {"success": False, "error": kb_result["error"]}

    kb_id = kb_result.get("id")
    print(f"  KB created: {kb_id} - {kb_result.get('name')}")

    # 2. Upload files
    all_files = list(table_files)

    # Copy existing athlete_events.xlsx if available
    if copy_existing_excel and EXISTING_EXCEL.exists():
        dest = OUTPUT_DIR / "athlete_events.xlsx"
        if not dest.exists():
            shutil.copy2(EXISTING_EXCEL, dest)
            print(f"  Copied existing test file: {dest.name} ({dest.stat().st_size / 1024 / 1024:.1f} MB)")
        all_files.append(dest)

    print(f"\n[2/3] Uploading {len(all_files)} files...")
    upload_results = []
    for i, fp in enumerate(all_files):
        print(f"  [{i + 1}/{len(all_files)}] Uploading {fp.name}...", end=" ")
        result = upload_file(kb_id, fp, folder="tables")
        if "error" in result:
            print(f"ERROR: {result['error']}")
        else:
            doc_id = result.get("id", "unknown")
            print(f"OK (doc_id: {doc_id})")
            upload_results.append({
                "file": fp.name,
                "doc_id": doc_id,
                "result": result,
            })

    # 3. Wait for processing
    print(f"\n[3/3] Waiting for processing...")
    docs = wait_for_documents_ready(kb_id, len(upload_results))

    # Summarize
    summary = {
        "kb_id": kb_id,
        "kb_name": kb_result.get("name"),
        "total_files": len(all_files),
        "upload_results": upload_results,
        "document_statuses": {
            d.get("originalName", d.get("id", "unknown")): d.get("status")
            for d in docs
        },
    }

    # Save summary
    summary_path = OUTPUT_DIR / "upload_summary.json"
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2, default=str)
    print(f"\n  Summary saved to: {summary_path}")

    return summary


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Table Test Data Generator + Upload")
    parser.add_argument("--upload", action="store_true", help="Generate files + create KB + upload")
    parser.add_argument("--upload-only", action="store_true", help="Skip generation, only create KB + upload")
    parser.add_argument("--skip-existing-excel", action="store_true", help="Don't copy athlete_events.xlsx")
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    generators = [
        ("employee_basic.csv", generate_employee_basic),
        ("sales_monthly.csv", generate_sales_monthly),
        ("products.csv", generate_products),
        ("inventory.xlsx", generate_inventory_xlsx),
        ("financial_report.xlsx", generate_financial_report_xlsx),
        ("department_hierarchy.xlsx", generate_department_hierarchy_xlsx),
        ("sensor_data.csv", generate_sensor_data),
        ("empty_table.csv", generate_empty_table),
        ("single_row.csv", generate_single_row),
        ("wide_table.csv", generate_wide_table),
        ("tab_separated.csv", generate_tab_separated),
        ("unicode_data.csv", generate_unicode_data),
        ("crossref_orders.xlsx", generate_crossref_orders_xlsx),
    ]

    if not args.upload_only:
        print("=" * 60)
        print("Generating test table files")
        print("=" * 60)

        generated_files = []
        for name, gen_func in generators:
            path = OUTPUT_DIR / name
            print(f"  Generating {name}...", end=" ")
            try:
                result_path = gen_func()
                size_kb = result_path.stat().st_size / 1024
                print(f"OK ({size_kb:.1f} KB)")
                generated_files.append(result_path)
            except Exception as e:
                print(f"ERROR: {e}")

        print(f"\nGenerated {len(generated_files)} files in {OUTPUT_DIR}")
    else:
        # Collect existing files
        generated_files = [OUTPUT_DIR / name for name, _ in generators if (OUTPUT_DIR / name).exists()]
        print(f"Found {len(generated_files)} existing files in {OUTPUT_DIR}")

    if args.upload or args.upload_only:
        print("\n" + "=" * 60)
        print("Creating KB and uploading files")
        print("=" * 60)

        summary = create_kb_and_upload(
            generated_files,
            copy_existing_excel=not args.skip_existing_excel,
        )

        if summary.get("kb_id"):
            # Print KB_IDS entry for run_benchmarks.py
            print(f"\n# Add to run_benchmarks.py KB_IDS:")
            print(f'    "表格处理测试库": "{summary["kb_id"]}",')

    print("\nDone!")


if __name__ == "__main__":
    main()
